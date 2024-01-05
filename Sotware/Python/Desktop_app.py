
import socket
from PyQt5.QtWidgets import *
from PyQt5 import uic
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication

class my_windows(QMainWindow):
    def __init__(self):
        print(socket.gethostbyname(socket.gethostname()))   #Ausgabe die aktuellen IP-Adresse
        super(my_windows, self).__init__()
        uic.loadUi("Windows_1.ui", self)                    #Aufruf des ersten Fensters
        self.show()                                         #Anzeige des ersten Fensters

        self.pushButton.clicked.connect(self.user_list)                         #Definition für verschiede Knöpfe für verschieden Funktionen
        self.pushButton_2.clicked.connect(self.user_create)                     #Definition für verschiede Knöpfe für verschieden Funktionen
        self.pushButton_3.clicked.connect(self.start_weighting)                 #Definition für verschiede Knöpfe für verschieden Funktionen
        self.pushButton_4.clicked.connect(lambda : self.compare_weighting(2))   #Definition für verschiede Knöpfe für verschieden Funktionen

    def compare_weighting(self, number):                #Die Funktion für Anzeige von Gewicht
        self.close()                                    #Das letzte Fenster schließen
        uic.loadUi("compare_weighting.ui", self)        #Aufruf von neuem Fenster
        people=[]
        wb = load_workbook('Data.xlsx')                 #Aufruf von Data.xlsx
        ws = wb.active
        for row in range(1,20):                         #Speichern der Personen in einem List names People
            char=get_column_letter(1)
            people.append(ws[char+str(row)].value)

        row=0
        self.tableWidget.setRowCount(len(people))
        for person in people:                           #Anzeige des namen der Personen auf dem Fenster
            self.tableWidget.setItem(row, 0, QTableWidgetItem(person))
            row=row+1

        char=get_column_letter(3)                               #Nimm den Inhalt von 3.Spalte
        self.lcdNumber.display(ws[char + str(number)].value)    #Anzeige den Inhalt von 3.Spalte

        char_1 = get_column_letter(5)                           #Nimm den Inhalt von 5.Spalte
        self.lcdNumber_2.display(ws[char_1+str(number)].value)  #Anzeige den Inhalt von 5.Spalte

        char_2=get_column_letter(4)                             #Nimm den Inhalt von 4.Spalte
        weight=int(ws[char_1+str(number)].value)                #Nimm das Gewicht von der Person
        height = int(ws[char_2+str(number)].value)              #Nimm die Größe von der Person
        height=height / 100                                     #Umwandeln das Gewicht in Meter
        print(height)
        bmi = weight / pow(height,2)                            #Berechnen von BMI
        self.lcdNumber_3.display(bmi)                           #Anzeige von BMI
        self.show()
        self.pushButton.clicked.connect(self.show_weight)       #Aufruf der nächste Funktion




    def show_weight(self):
        row = self.spinBox.value()                              #Nimm den Inhalt von dem Box
        wb = load_workbook('Data.xlsx')                         #Aufruf von dem Data.xlsx
        ws = wb.active

        #Die Unten stehenden Zeilen sind wie vorherige Funktion
        char=get_column_letter(3)
        self.lcdNumber.display(ws[char + str(row)].value)

        char_1 = get_column_letter(5)
        self.lcdNumber_2.display(ws[char_1+str(row)].value)

        char_2=get_column_letter(4)
        weight=int(ws[char_1+str(row)].value)
        height = int(ws[char_2+str(row)].value)
        height=height / 100
        print(height)
        bmi = weight / pow(height,2)
        self.lcdNumber_3.display(bmi)
    def start_weighting(self):                  #Die Funktion für Start die Verbindung zwischen ESP32 und Software für Erhalt des Gewichts
        row = int(self.lineEdit.text())         #Eingeben die Nummer der Personen(Die Zeile von Excel Datei)
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        char = get_column_letter(5)

        PORT = 1234                             #Definition von einem beliebigen Port

        #Konfiguratin von der Verbindung
        SERVER = socket.gethostbyname(socket.gethostname())     #Nimm die IP-Adresse
        ADDR = (SERVER, PORT)                                   #Erstellen einer Verbindung mit dem Port und IP-Adresse
        client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        client_socket.bind(ADDR)                                #Verbindung mit dem Client

        client_socket.listen(1)                                 #Start von der Verbindung
        total_weight = 0
        repeat = 0
        while(repeat<5):                                        #Nimm das Gewicht 5 Mal
            conn, addr = client_socket.accept()
            data = conn.recv(1024)
            if data:
                message = data.decode().strip()                 #Erhalten und Entschlüsseln von dem Data
                message = float(message)
                total_weight += message
                repeat += 1

            else:
                conn.close()
                continue
        conn.close()                                            #Scließen der Verbindung

        ws[char + str(row)].value = (total_weight/5)            #Speichern das Gewicht für Personen, Hier wird der Durschnitt von dem Gewicht gespeichert
        wb.save('Data.xlsx')
        self.compare_weighting(row)

    #Erstellen für neuen Benutzer
    def user_create(self):
        self.close()
        uic.loadUi("untitled.ui", self)
        self.show()
        self.pushButton.clicked.connect(self.data)

    #Anzeige von den Benutzern
    def user_list(self):
        people=[]
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        for row in range(1,20):
            char=get_column_letter(1)
            people.append(ws[char+str(row)].value)

        row=0
        self.tableWidget.setRowCount(len(people))
        for person in people:
            self.tableWidget.setItem(row, 0, QTableWidgetItem(person))
            row=row+1

    #Erstellen von neuen Benutzern
    def data(self):

        wb = load_workbook('Data.xlsx')
        ws = wb.active

        ws.append([self.lineEdit.text(),self.lineEdit_2.text(),self.lineEdit_3.text(),self.lineEdit_4.text()])  #Speichern des Datei von dem Benutzer
        wb.save('Data.xlsx')

        self.lineEdit.clear()
        self.lineEdit_2.clear()
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.close()
        uic.loadUi("Windows_1.ui", self)
        self.show()
        self.pushButton.clicked.connect(self.user_list)
        self.pushButton_2.clicked.connect(self.user_create)
        self.pushButton_3.clicked.connect(self.start_weighting)
        self.pushButton_4.clicked.connect(self.compare_weighting)


def window():
    app= QApplication([])
    win= my_windows()
    win.show()
    app.exec_()

window()